"""
Excel Exporter with LRIS Schema Compliance

Generates formatted Excel files with conditional formatting and validation.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
from typing import List, Dict, Any
from datetime import datetime

from src.core.logger import get_logger
from src.core.config import get_config
from src.core.exceptions import ExcelExportError
from models import JamabandiRecord, VillageJamabandi

logger = get_logger(__name__)


class ExcelExporter:
    """
    Export Jamabandi records to LRIS-compliant Excel format
    """
    
    # LRIS schema column order
    COLUMNS = [
        'number_khevat',
        'number_khata',
        'nam_tarf_ya_patti',
        'nam_malik_meh_ahval',
        'cultivator_name',
        'cultivator_parentage',
        'cultivator_caste',
        'cultivator_village',
        'cultivator_ownership',
        'vasayil_abapashi',
        'khasra_hal',
        'khasra_sabik',
        'area_kanal',
        'area_marla',
        'kisam_zamin',
        'lagan_details',
        'mutalba_details',
        'havala_intakal',
        'kaifiyat',
        'extraction_confidence',
        'extraction_method',
        'needs_review',
        'page_number',
    ]
    
    def __init__(self):
        """Initialize exporter"""
        self.config = get_config().output
        logger.info("Excel exporter initialized")
    
    def export(
        self,
        village_data: VillageJamabandi,
        output_path: Path
    ) -> Path:
        """
        Export village data to Excel
        
        Args:
            village_data: Complete village Jamabandi data
            output_path: Output file path
            
        Returns:
            Path to created Excel file
        """
        try:
            logger.info(f"Exporting {len(village_data.records)} records to {output_path}")
            
            # Ensure output directory exists
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            # Convert records to DataFrame
            df = self._records_to_dataframe(village_data.records)
            
            # Create Excel workbook
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Write main data sheet
                df.to_excel(writer, sheet_name='Land Records', index=False)
                
                # Write summary sheet
                summary_df = self._create_summary(village_data)
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                # Write metadata sheet
                metadata_df = self._create_metadata(village_data)
                metadata_df.to_excel(writer, sheet_name='Metadata', index=False)
                
                # Apply formatting
                workbook = writer.book
                self._apply_formatting(workbook, village_data)
            
            logger.info(f"Successfully exported to {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Excel export failed: {e}")
            raise ExcelExportError(str(output_path), str(e))
    
    def _records_to_dataframe(self, records: List[JamabandiRecord]) -> pd.DataFrame:
        """Convert records to DataFrame"""
        data = [record.to_flat_dict() for record in records]
        df = pd.DataFrame(data)
        
        # Ensure all columns exist
        for col in self.COLUMNS:
            if col not in df.columns:
                df[col] = None
        
        # Reorder columns
        df = df[self.COLUMNS]
        
        return df
    
    def _create_summary(self, village_data: VillageJamabandi) -> pd.DataFrame:
        """Create summary statistics"""
        summary = village_data.get_summary()
        
        summary_data = [
            {'Metric': 'Village', 'Value': village_data.metadata.babata_mouza},
            {'Metric': 'District', 'Value': village_data.metadata.zla},
            {'Metric': 'Tehsil', 'Value': village_data.metadata.thsil},
            {'Metric': 'Year', 'Value': village_data.metadata.sal},
            {'Metric': '', 'Value': ''},
            {'Metric': 'Total Records', 'Value': summary['total_records']},
            {'Metric': 'Total Pages', 'Value': village_data.total_pages},
            {'Metric': 'Unique Khata', 'Value': summary.get('unique_khata', 0)},
            {'Metric': 'Unique Khasra', 'Value': summary.get('unique_khasra', 0)},
            {'Metric': '', 'Value': ''},
            {'Metric': 'Average Confidence', 'Value': f"{summary['average_confidence']:.1%}"},
            {'Metric': 'Records Needing Review', 'Value': summary['records_needing_review']},
            {'Metric': 'Review Percentage', 'Value': f"{summary['review_percentage']:.1f}%"},
            {'Metric': '', 'Value': ''},
            {'Metric': 'Processed At', 'Value': village_data.processed_at.strftime('%Y-%m-%d %H:%M:%S')},
            {'Metric': 'Source File', 'Value': village_data.source_file},
        ]
        
        return pd.DataFrame(summary_data)
    
    def _create_metadata(self, village_data: VillageJamabandi) -> pd.DataFrame:
        """Create metadata sheet"""
        metadata_data = [
            {'Field': 'Application', 'Value': 'AgriStack Land Record Digitization'},
            {'Field': 'Version', 'Value': '1.0.0'},
            {'Field': 'Schema', 'Value': 'LRIS Compliant'},
            {'Field': '', 'Value': ''},
            {'Field': 'Village', 'Value': village_data.metadata.babata_mouza},
            {'Field': 'District (Zla)', 'Value': village_data.metadata.zla},
            {'Field': 'Tehsil', 'Value': village_data.metadata.thsil},
            {'Field': 'Year (Sal)', 'Value': village_data.metadata.sal},
            {'Field': '', 'Value': ''},
            {'Field': 'Source PDF', 'Value': village_data.source_file},
            {'Field': 'Processed Date', 'Value': village_data.processed_at.strftime('%Y-%m-%d')},
            {'Field': 'Processed Time', 'Value': village_data.processed_at.strftime('%H:%M:%S')},
        ]
        
        return pd.DataFrame(metadata_data)
    
    def _apply_formatting(self, workbook: Workbook, village_data: VillageJamabandi):
        """Apply Excel formatting"""
        # Format Land Records sheet
        if 'Land Records' in workbook.sheetnames:
            ws = workbook['Land Records']
            self._format_data_sheet(ws, village_data)
        
        # Format Summary sheet
        if 'Summary' in workbook.sheetnames:
            ws = workbook['Summary']
            self._format_summary_sheet(ws)
        
        # Format Metadata sheet
        if 'Metadata' in workbook.sheetnames:
            ws = workbook['Metadata']
            self._format_metadata_sheet(ws)
    
    def _format_data_sheet(self, ws, village_data: VillageJamabandi):
        """Format main data sheet"""
        # Header formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Highlight low confidence rows
        if self.config.highlight_low_confidence:
            yellow_fill = PatternFill(
                start_color=self.config.highlight_color,
                end_color=self.config.highlight_color,
                fill_type="solid"
            )
            
            # Find needs_review column
            needs_review_col = None
            for idx, cell in enumerate(ws[1], 1):
                if cell.value == 'needs_review':
                    needs_review_col = idx
                    break
            
            if needs_review_col:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    if row[needs_review_col - 1].value == True:
                        for cell in row:
                            cell.fill = yellow_fill
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    def _format_summary_sheet(self, ws):
        """Format summary sheet"""
        # Bold metric names
        for row in ws.iter_rows(min_row=2):
            row[0].font = Font(bold=True)
        
        # Auto-adjust widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40
    
    def _format_metadata_sheet(self, ws):
        """Format metadata sheet"""
        # Bold field names
        for row in ws.iter_rows(min_row=2):
            row[0].font = Font(bold=True)
        
        # Auto-adjust widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
